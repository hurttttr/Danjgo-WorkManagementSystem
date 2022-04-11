import os.path
import shutil
import openpyxl
import pytz
import smtplib
import config
from datetime import datetime
from openpyxl.styles import PatternFill
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from django import forms
from django.contrib.auth.models import User
from app import models


# 创建文件夹
def createDirectory(path):
    if not os.path.exists(path):
        os.makedirs(path)


# 删除文件夹以及压缩包
def deleteDirectory(path):
    if os.path.exists(path):  # 删除文件夹
        shutil.rmtree(path)
    zip_path = path + '.zip'
    if os.path.exists(zip_path):  # 删除压缩包
        shutil.rmtree(zip_path)


# 保存文件
def saveUploadedFile(f, path):
    with open(path, 'wb') as destination:
        for chunk in f.chunks():
            destination.write(chunk)


# 填写完成记录
def writeRecord(path, id):
    user_list = models.Upload.objects.filter(work_id=id)
    lst = []
    for i in user_list:
        lst.append(i.username)
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    for index, value in enumerate(sheet.rows, 1):
        if index > 1:
            user_id = list(map(lambda x: x.value, value))[0]
            if user_id in lst:
                sheet.cell(row=index, column=3).value = '完成'
            else:
                sheet.cell(row=index, column=3).value = '未完成'
                sheet.cell(row=index, column=3).fill = PatternFill(fill_type='solid', fgColor="FFC125")
    wb.save(path)
    shutil.copy(path, 'upload/{}/完成情况.xlsx'.format(id))


# 压缩文件夹
def zipDirectory(id):
    path = 'upload/' + str(id)
    shutil.make_archive(path, 'zip', path)


# 导入名单
def addUsers():
    path = 'upload/cache/名单.xlsx'
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    for index, value in enumerate(sheet.rows, 1):
        if index > 1:
            lst = list(map(lambda x: x.value, value))
            user = User.objects.filter(username=lst[0])  # 防止重复导入
            if not user:
                if lst[2] != 1:
                    User.objects.create_user(username=lst[0], password='123456', first_name=lst[1])
                else:
                    User.objects.create_superuser(username=lst[0], password='123456', first_name=lst[1])


# 检查作业过期情况
def checkDeadline():
    work_list = models.Work.objects.all()
    nowtime = datetime.now().replace(tzinfo=pytz.UTC)
    nowtime.tzinfo == None
    for i in work_list:
        if nowtime > i.deadline:  # 已过期
            if not i.send:
                if i.email:
                    try:
                        shutil.copy('upload/cache/名单.xlsx', 'upload/{}/完成情况.xlsx'.format(i.id))  # 拷贝名单
                        writeRecord('upload/{}/完成情况.xlsx'.format(i.id), i.id)  # 写入完成情况
                        path = 'upload/' + str(i.id)
                        shutil.make_archive(path, 'zip', path)  # 压缩文件夹
                        senfEmail(i.email, i.name, 'upload/{}.zip'.format(i.id))  # 发送邮件
                        print('已发送作业：{}'.format(i.name))
                        i.send = True
                        i.save()
                    except Exception as e:
                        print(e)


# 发送邮件
def senfEmail(emailAddress, filename, filepath):
    my_mail = config.my_mail
    key = config.key

    smtp_server = 'smtp.qq.com'

    server = smtplib.SMTP(smtp_server)
    server.connect(smtp_server, 25)
    server.login(my_mail, key)

    msg = MIMEMultipart()
    msg['From'] = Header('作业管理系统')  # 发件人
    msg['To'] = Header(emailAddress)  # 收件人
    msg['Subject'] = Header(filename)  # 邮件标题

    part = MIMEApplication(open(filepath, 'rb').read())
    part.add_header('Content-Disposition', 'attachment',
                    filename=('utf-8', '', filename + '.zip'))
    msg.attach(part)
    server.sendmail(my_mail, emailAddress, msg.as_string())
    server.close()


# 文件类
class UploadFileForm(forms.Form):
    title = forms.CharField(max_length=50)
    file = forms.FileField()
